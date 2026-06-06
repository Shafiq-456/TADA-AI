"""
Report Generator Service - Compiles PDF reports using ReportLab
"""
import os
import json
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, Optional

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

class ReportGenerator:
    """Generates professional analytics reports for datasets"""

    def generate_pdf_report(self, dataset_id: str, title: str, meta: Dict[str, Any], analytics: Optional[Dict[str, Any]] = None) -> str:
        """Create a PDF report and return the file path"""
        report_dir = Path("uploads") / dataset_id / "reports"
        report_dir.mkdir(parents=True, exist_ok=True)
        
        filename = f"report_{int(datetime.utcnow().timestamp())}.pdf"
        file_path = report_dir / filename
        
        doc = SimpleDocTemplate(
            str(file_path),
            pagesize=letter,
            rightMargin=54,
            leftMargin=54,
            topMargin=54,
            bottomMargin=54
        )
        
        styles = getSampleStyleSheet()
        
        # Custom premium styling
        primary_color = colors.HexColor("#2563EB")
        secondary_color = colors.HexColor("#0F172A")
        accent_color = colors.HexColor("#10B981")
        text_color = colors.HexColor("#334155")
        
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=24,
            textColor=secondary_color,
            spaceAfter=15
        )
        
        subtitle_style = ParagraphStyle(
            'ReportSubtitle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=11,
            textColor=primary_color,
            spaceAfter=25
        )
        
        h1_style = ParagraphStyle(
            'H1Style',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=16,
            textColor=secondary_color,
            spaceBefore=15,
            spaceAfter=10
        )
        
        body_style = ParagraphStyle(
            'BodyStyle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            textColor=text_color,
            leading=14,
            spaceAfter=10
        )
        
        bold_body_style = ParagraphStyle(
            'BoldBodyStyle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=10,
            textColor=secondary_color,
            leading=14,
            spaceAfter=10
        )
        
        story = []
        
        # Title and Subtitle
        story.append(Paragraph(title, title_style))
        story.append(Paragraph(f"Generated automatically by TADA AI on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", subtitle_style))
        story.append(Spacer(1, 15))
        
        # Executive Summary Section
        story.append(Paragraph("1. Executive Summary", h1_style))
        summary_text = (
            f"This executive report details the analytics profiles for the dataset <b>{meta.get('name')}</b>. "
            f"The dataset consists of <b>{meta.get('row_count', 'unknown'):,}</b> rows and <b>{meta.get('column_count', 'unknown')}</b> columns. "
            f"Overall data quality health was evaluated at <b>{meta.get('quality_score', 'unknown')}%</b>."
        )
        story.append(Paragraph(summary_text, body_style))
        story.append(Spacer(1, 10))
        
        # Metadata / Stats Table
        story.append(Paragraph("2. Dataset Summary Metrics", h1_style))
        
        kpis = [
            ["Metric", "Value"],
            ["Dataset Name", meta.get("name", "Unknown")],
            ["File Size", f"{round(meta.get('file_size', 0) / 1024, 2)} KB"],
            ["Total Rows", f"{meta.get('row_count', 0):,}"],
            ["Total Columns", str(meta.get('column_count', 0))],
            ["Quality Score", f"{meta.get('quality_score', 0)}%"],
            ["Missing Cells", f"{meta.get('missing_values', 0)} cells"],
            ["Duplicate Rows", f"{meta.get('duplicate_rows', 0)} rows"]
        ]
        
        t = Table(kpis, colWidths=[200, 300])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (1, 0), primary_color),
            ('TEXTCOLOR', (0, 0), (1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ('PADDING', (0, 0), (-1, -1), 6),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        story.append(t)
        story.append(Spacer(1, 15))
        
        # Schema Section
        story.append(Paragraph("3. Data Schema & Architecture", h1_style))
        story.append(Paragraph("A breakdown of columns, data types, null occurrences, and unique cardinality counts:", body_style))
        
        schema_rows = [["Column Name", "Type", "Null Count", "Unique Values"]]
        for col in meta.get("columns", [])[:15]:  # Limit to top 15 to avoid massive overflow
            schema_rows.append([
                col.get("name", ""),
                col.get("type", ""),
                f"{col.get('null_count', 0)} ({col.get('null_pct', 0)}%)",
                str(col.get("unique_count", 0))
            ])
            
        t_schema = Table(schema_rows, colWidths=[150, 100, 120, 130])
        t_schema.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), secondary_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ('PADDING', (0, 0), (-1, -1), 5),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]))
        story.append(t_schema)
        story.append(Spacer(1, 15))

        # Insights Section
        story.append(Paragraph("4. Auto-Generated Insights & Quality Report", h1_style))
        
        issues = []
        if meta.get("missing_values", 0) > 0:
            issues.append(f"• Found <b>{meta.get('missing_values')}</b> missing values in the dataset. Missing values can skew statistical calculations and should be filled or resolved.")
        else:
            issues.append("• <b>Completeness</b>: Excellent! No missing cells detected.")
            
        if meta.get("duplicate_rows", 0) > 0:
            issues.append(f"• Found <b>{meta.get('duplicate_rows')}</b> duplicate rows. Standard practice suggests eliminating exact duplicates to avoid over-weighting specific records.")
        else:
            issues.append("• <b>Uniqueness</b>: Excellent! No duplicated rows found.")

        for col in meta.get("columns", []):
            if col.get("null_pct", 0) > 10:
                issues.append(f"• Column <b>'{col['name']}'</b> has a high percentage of missing cells ({col['null_pct']}%). This field might need to be dropped or heavily cleaned before building predictive models.")
        
        if len(issues) == 0:
            issues.append("• No major data quality anomalies found. The dataset is clean and ready for machine learning modelling.")
            
        for issue in issues:
            story.append(Paragraph(issue, body_style))
            
        doc.build(story)
        return str(file_path)

    def generate_excel_report(self, dataset_id: str, title: str, meta: Dict[str, Any]) -> str:
        """Create an Excel report (.xlsx) and return the file path"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        report_dir = Path("uploads") / dataset_id / "reports"
        report_dir.mkdir(parents=True, exist_ok=True)
        
        filename = f"report_{int(datetime.utcnow().timestamp())}.xlsx"
        file_path = report_dir / filename
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Executive Summary"
        
        # Enable grid lines
        ws.views.sheetView[0].showGridLines = True
        
        # Styles
        title_font = Font(name="Calibri", size=18, bold=True, color="1F497D")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        section_font = Font(name="Calibri", size=14, bold=True, color="1F497D")
        normal_font = Font(name="Calibri", size=11)
        
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        # Write Title
        ws["A1"] = title
        ws["A1"].font = title_font
        ws.row_dimensions[1].height = 30
        
        ws["A2"] = f"Generated by TADA AI on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = Font(name="Calibri", size=10, italic=True)
        
        # 1. Summary Metrics
        ws["A4"] = "Dataset Summary Metrics"
        ws["A4"].font = section_font
        
        headers = ["Metric", "Value"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")
            cell.border = thin_border
            
        kpis = [
            ["Dataset Name", meta.get("name", "Unknown")],
            ["File Size", f"{round(meta.get('file_size', 0) / 1024, 2)} KB"],
            ["Total Rows", meta.get('row_count', 0)],
            ["Total Columns", meta.get('column_count', 0)],
            ["Quality Score", f"{meta.get('quality_score', 0)}%"],
            ["Missing Cells", meta.get('missing_values', 0)],
            ["Duplicate Rows", meta.get('duplicate_rows', 0)]
        ]
        
        for r_idx, row in enumerate(kpis, 6):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = normal_font
                cell.border = thin_border
                if r_idx % 2 == 1:
                    cell.fill = zebra_fill
                    
        # 2. Columns sheet
        ws_schema = wb.create_sheet(title="Data Schema")
        ws_schema.views.sheetView[0].showGridLines = True
        
        ws_schema["A1"] = "Data Schema & Architecture"
        ws_schema["A1"].font = title_font
        
        schema_headers = ["Column Name", "Type", "Null Count", "Null Pct", "Unique Count", "Unique Pct"]
        for col_num, header in enumerate(schema_headers, 1):
            cell = ws_schema.cell(row=3, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")
            cell.border = thin_border
            
        for r_idx, col in enumerate(meta.get("columns", []), 4):
            row_data = [
                col.get("name", ""),
                col.get("type", ""),
                col.get("null_count", 0),
                f"{col.get('null_pct', 0)}%",
                col.get("unique_count", 0),
                f"{col.get('unique_pct', 0)}%"
            ]
            for c_idx, val in enumerate(row_data, 1):
                cell = ws_schema.cell(row=r_idx, column=c_idx, value=val)
                cell.font = normal_font
                cell.border = thin_border
                if r_idx % 2 == 1:
                    cell.fill = zebra_fill
                    
        # Auto-adjust column widths
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        wb.save(file_path)
        return str(file_path)
